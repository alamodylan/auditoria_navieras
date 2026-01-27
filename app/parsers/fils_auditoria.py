# app/parsers/fils_auditoria.py

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from typing import Dict, List, Optional, Tuple, Iterable, Any

import unicodedata
from openpyxl import load_workbook

from app.parsers.base import BaseParser
from app.parsers.normalization import normalize_guia, normalize_contenedor
from app.utils.money import parse_money
from app.utils.logging import get_logger

logger = get_logger("parser.fils_auditoria")


# =========================
# Helpers de normalización
# =========================

def _strip_accents(s: str) -> str:
    if s is None:
        return ""
    return "".join(
        c for c in unicodedata.normalize("NFKD", str(s))
        if not unicodedata.combining(c)
    )

def _norm_header(s: Any) -> str:
    """
    Normaliza encabezados para hacer matching robusto:
    - lower
    - sin acentos
    - espacios colapsados
    """
    if s is None:
        return ""
    s = _strip_accents(str(s)).strip().lower()
    s = " ".join(s.split())
    return s

def _parse_fecha(value) -> Optional[datetime]:
    """
    Parsea fecha del FILS:
    - openpyxl puede traer datetime
    - o string tipo: 01/10/2025 08:49
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    s = str(value).strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%d/%m/%Y %H:%M")
    except Exception:
        return None


def _find_header_row(
    ws,
    required_tokens: List[str],
    max_scan_rows: int = 30
) -> Tuple[int, List[str], List[str]]:
    """
    Busca una fila que parezca encabezado.
    Retorna: (header_row_idx_1based, headers_norm, warnings)
    """
    warnings: List[str] = []
    required = [_norm_header(t) for t in required_tokens if t]

    # Escanear primeras filas para detectar encabezado
    rows = ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True)
    candidate_idx = None
    candidate_headers = None

    for i, row in enumerate(rows, start=1):
        headers_norm = [_norm_header(x) for x in (row or [])]
        joined = " | ".join(headers_norm)
        hits = sum(1 for t in required if t and t in joined)

        # Si pega todos los tokens, lo damos por header
        if hits >= max(1, len(required)):
            candidate_idx = i
            candidate_headers = headers_norm
            break

        # Heurística: si al menos 2 tokens aparecen, guardamos como candidato
        if hits >= 2 and candidate_idx is None:
            candidate_idx = i
            candidate_headers = headers_norm

    if candidate_idx is None or candidate_headers is None:
        # fallback
        warnings.append("Encabezado no detectado claramente; usando fila 1 como encabezado.")
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or []
        return 1, [_norm_header(x) for x in row1], warnings

    if candidate_idx != 1:
        warnings.append(f"Encabezado no detectado claramente en la fila 1; usando fila {candidate_idx} como encabezado.")
    return candidate_idx, candidate_headers, warnings


def _map_idx(headers_norm: List[str], synonyms: Dict[str, List[str]]) -> Dict[str, Optional[int]]:
    """
    headers_norm: encabezados ya normalizados (_norm_header)
    synonyms: {canonical: [sin1, sin2...]} (también normalizados internamente)
    Retorna {canonical: idx or None}
    """
    out: Dict[str, Optional[int]] = {}
    for canon, opts in synonyms.items():
        idx = None
        for opt in opts:
            optn = _norm_header(opt)
            # match exact
            for i, hn in enumerate(headers_norm):
                if hn == optn:
                    idx = i
                    break
            if idx is not None:
                break
            # match contains
            for i, hn in enumerate(headers_norm):
                if optn and optn in hn:
                    idx = i
                    break
            if idx is not None:
                break
        out[canon] = idx
    return out


def _cell(row: Tuple, idx: Optional[int]):
    if idx is None:
        return None
    if row is None:
        return None
    return row[idx] if idx < len(row) else None


# =========================
# Parser
# =========================

@dataclass
class _CargoEvent:
    guia: str
    cargo_key: str
    fecha: Optional[datetime]
    accion: str
    cargo_id: str
    cargo_nombre: str
    moneda: str
    monto: Decimal


class FILSAuditoriaParser(BaseParser):
    """
    Este parser une 3 hojas del FILS:
      - "Guía"                 -> base (estado, ruta, monto tarifa, etc.)
      - "Contenedor"           -> guía -> contenedor (necesario para ONE cuando no trae guía)
      - "Cargos Adicionales"   -> cargos naviera por guía (solo el último evento; si termina en Eliminar, se descarta)

    Output (por guía):
      {
        guia, contenedor, estado, fecha, fecha_cierre,
        monto_total (Monto Tarifa), ruta,
        cargos: [{cargo_id, cargo, moneda, monto}]
      }
    """

    # Nombres esperados de hojas (según tu archivo)
    SHEET_GUIA = "Guía"
    SHEET_CONT = "Contenedor"
    SHEET_CARGOS = "Cargos Adicionales"

    def sniff(self, path: str) -> Dict:
        meta = {
            "ok": True,
            "sheets": [],
            "warnings": [],
            "errors": [],
            "headers": {},
        }

        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            meta["ok"] = False
            meta["errors"].append(f"No se pudo abrir el Excel: {e}")
            return meta

        meta["sheets"] = wb.sheetnames

        def _sniff_sheet(sheet_name: str, required_tokens: List[str]):
            if sheet_name not in wb.sheetnames:
                meta["warnings"].append(f"No existe la hoja '{sheet_name}' en el archivo.")
                return
            ws = wb[sheet_name]
            hr, headers_norm, warns = _find_header_row(ws, required_tokens=required_tokens)
            meta["warnings"].extend(warns)
            meta["headers"][sheet_name] = {
                "header_row": hr,
                "headers_norm": headers_norm[:60],
            }

        _sniff_sheet(self.SHEET_GUIA, required_tokens=["numero guia", "fecha", "estado", "ruta", "monto tarifa"])
        _sniff_sheet(self.SHEET_CONT, required_tokens=["numero guia", "contenedor"])
        _sniff_sheet(self.SHEET_CARGOS, required_tokens=["numero guia", "cargo", "accion", "fecha", "total naviera"])

        try:
            wb.close()
        except Exception:
            pass

        return meta

    def parse(self, path: str) -> List[dict]:
        wb = load_workbook(path, read_only=True, data_only=True)

        # ---------
        # 1) Base: hoja Guía
        # ---------
        guia_map: Dict[str, dict] = {}

        if self.SHEET_GUIA not in wb.sheetnames:
            wb.close()
            raise ValueError(f"FILS: no existe la hoja '{self.SHEET_GUIA}'.")

        ws_g = wb[self.SHEET_GUIA]
        hr_g, headers_g, warns = _find_header_row(
            ws_g,
            required_tokens=["numero guia", "fecha", "estado", "ruta", "monto tarifa"]
        )
        for w in warns:
            logger.warning(w)

        syn_g = {
            "guia": ["número guía", "numero guia", "numero guía", "guia", "nro guia"],
            "estado": ["estado"],
            "fecha": ["fecha"],
            "ruta": ["ruta"],
            "monto_tarifa": ["monto tarifa", "monto_tarif", "tarifa", "monto total", "total"],
        }
        idx_g = _map_idx(headers_g, syn_g)

        if idx_g["guia"] is None:
            wb.close()
            raise ValueError("FILS/Guía: no se encontró columna 'Número Guía'.")
        if idx_g["monto_tarifa"] is None:
            wb.close()
            raise ValueError("FILS/Guía: no se encontró columna 'Monto Tarifa' (total a comparar).")

        logger.info(f"FILS/Guía idx={idx_g} header_row={hr_g}")

        for row in ws_g.iter_rows(min_row=hr_g + 1, values_only=True):
            g = normalize_guia(_cell(row, idx_g["guia"]))
            if not g:
                continue

            estado = str(_cell(row, idx_g["estado"]) or "").strip().upper()
            fecha = _parse_fecha(_cell(row, idx_g["fecha"]))
            ruta = str(_cell(row, idx_g["ruta"]) or "").strip()

            monto_tarifa_raw = _cell(row, idx_g["monto_tarifa"])
            monto_tarifa = parse_money(monto_tarifa_raw)

            # quedarnos con el evento más reciente (para campos base)
            prev = guia_map.get(g)
            if prev is None:
                guia_map[g] = {
                    "guia": g,
                    "contenedor": "",
                    "estado": estado,
                    "fecha": fecha,
                    "fecha_cierre": fecha,
                    "ruta": ruta,
                    "monto_total": monto_tarifa,
                    "monto_flete": None,
                    "monto_extras": None,
                    "cargos": [],
                }
            else:
                # decidir por "fecha" más reciente
                prev_fecha = prev.get("fecha") or prev.get("fecha_cierre")
                if (fecha is not None) and (prev_fecha is None or fecha >= prev_fecha):
                    prev["estado"] = estado or prev.get("estado", "")
                    prev["fecha"] = fecha
                    prev["fecha_cierre"] = fecha
                    prev["ruta"] = ruta or prev.get("ruta", "")
                    prev["monto_total"] = monto_tarifa if monto_tarifa is not None else prev.get("monto_total", Decimal("0"))

        # ---------
        # 2) Hoja Contenedor: guia -> contenedor
        # ---------
        if self.SHEET_CONT in wb.sheetnames:
            ws_c = wb[self.SHEET_CONT]
            hr_c, headers_c, warns = _find_header_row(ws_c, required_tokens=["numero guia", "contenedor"])
            for w in warns:
                logger.warning(w)

            syn_c = {
                "guia": ["número guía", "numero guia", "guia"],
                "contenedor": ["contenedor"],
                "fecha": ["fecha"],
                "accion": ["accion", "acción"],
            }
            idx_c = _map_idx(headers_c, syn_c)

            if idx_c["guia"] is None:
                wb.close()
                raise ValueError("FILS/Contenedor: no se encontró columna 'Número Guía'.")
            if idx_c["contenedor"] is None:
                # aquí NO lo tiramos como error fatal: hay casos donde el contenedor viene en otra columna rara,
                # pero para tu formato debería existir. Lo dejamos como error explícito.
                wb.close()
                raise ValueError("FILS/Contenedor: no se encontró columna 'Contenedor'.")

            logger.info(f"FILS/Contenedor idx={idx_c} header_row={hr_c}")

            # nos quedamos con el último contenedor por guía (por fecha si existe)
            cont_last: Dict[str, Tuple[Optional[datetime], str]] = {}

            for row in ws_c.iter_rows(min_row=hr_c + 1, values_only=True):
                g = normalize_guia(_cell(row, idx_c["guia"]))
                if not g:
                    continue

                cont = normalize_contenedor(_cell(row, idx_c["contenedor"]))
                cont = cont.upper().replace("-", "").replace(" ", "")
                if not cont:
                    continue

                fecha = _parse_fecha(_cell(row, idx_c["fecha"])) if idx_c.get("fecha") is not None else None

                prev_fecha, _prev_cont = cont_last.get(g, (None, ""))
                if prev_fecha is None:
                    cont_last[g] = (fecha, cont)
                else:
                    # si no hay fecha nueva, conservamos la existente
                    if fecha is not None and fecha >= prev_fecha:
                        cont_last[g] = (fecha, cont)

            # inyectar en guia_map
            for g, (_f, cont) in cont_last.items():
                if g in guia_map:
                    guia_map[g]["contenedor"] = cont
                else:
                    # Si aparece guía solo en Contenedor (raro, pero posible), la creamos
                    guia_map[g] = {
                        "guia": g,
                        "contenedor": cont,
                        "estado": "",
                        "fecha": None,
                        "fecha_cierre": None,
                        "ruta": "",
                        "monto_total": Decimal("0"),
                        "monto_flete": None,
                        "monto_extras": None,
                        "cargos": [],
                    }

        else:
            logger.warning("FILS: no existe hoja 'Contenedor' (se continuará sin matching por contenedor).")

        # ---------
        # 3) Hoja Cargos Adicionales: últimos cargos por guía
        # ---------
        if self.SHEET_CARGOS in wb.sheetnames:
            ws_a = wb[self.SHEET_CARGOS]
            hr_a, headers_a, warns = _find_header_row(
                ws_a,
                required_tokens=["numero guia", "cargo", "accion", "fecha", "total naviera"]
            )
            for w in warns:
                logger.warning(w)

            syn_a = {
                "guia": ["número guía", "numero guia", "guia"],
                "accion": ["acción", "accion"],
                "fecha": ["fecha"],
                "cargo_id": ["cargo id", "id cargo", "cargoid"],
                "cargo": ["cargo"],
                "moneda_naviera": ["moneda naviera", "moneda"],
                "monto_naviera": ["total naviera", "monto naviera"],  # priorizamos TOTAL NAVIERA
            }
            idx_a = _map_idx(headers_a, syn_a)

            if idx_a["guia"] is None:
                wb.close()
                raise ValueError("FILS/Cargos Adicionales: no se encontró columna 'Número Guía'.")
            if idx_a["cargo"] is None and idx_a["cargo_id"] is None:
                wb.close()
                raise ValueError("FILS/Cargos Adicionales: no se encontró columna 'Cargo' o 'Cargo Id'.")
            if idx_a["monto_naviera"] is None:
                wb.close()
                raise ValueError("FILS/Cargos Adicionales: no se encontró columna 'Total Naviera' / 'Monto Naviera'.")

            logger.info(f"FILS/Cargos Adicionales idx={idx_a} header_row={hr_a}")

            # Guardar el último evento por (guia, cargo_key)
            last_event: Dict[Tuple[str, str], _CargoEvent] = {}

            for row in ws_a.iter_rows(min_row=hr_a + 1, values_only=True):
                g = normalize_guia(_cell(row, idx_a["guia"]))
                if not g:
                    continue

                accion = str(_cell(row, idx_a["accion"]) or "").strip().upper()
                fecha = _parse_fecha(_cell(row, idx_a["fecha"]))

                cargo_id = str(_cell(row, idx_a["cargo_id"]) or "").strip()
                cargo_nombre = str(_cell(row, idx_a["cargo"]) or "").strip()
                moneda = str(_cell(row, idx_a["moneda_naviera"]) or "").strip().upper()

                monto = parse_money(_cell(row, idx_a["monto_naviera"]))

                # definir key estable (preferimos cargo_id si existe)
                cargo_key = cargo_id.strip() if cargo_id.strip() else cargo_nombre.strip().upper()
                if not cargo_key:
                    continue

                key = (g, cargo_key)
                prev = last_event.get(key)

                ev = _CargoEvent(
                    guia=g,
                    cargo_key=cargo_key,
                    fecha=fecha,
                    accion=accion,
                    cargo_id=cargo_id,
                    cargo_nombre=cargo_nombre,
                    moneda=moneda,
                    monto=monto,
                )

                if prev is None:
                    last_event[key] = ev
                else:
                    # elegir por fecha más reciente; si no hay fecha, dejamos el existente
                    if prev.fecha is None and ev.fecha is not None:
                        last_event[key] = ev
                    elif ev.fecha is not None and prev.fecha is not None and ev.fecha >= prev.fecha:
                        last_event[key] = ev
                    # si fechas iguales o nulas, conservamos el último "visto"
                    elif ev.fecha is None and prev.fecha is None:
                        last_event[key] = ev

            # transformar eventos en cargos vigentes
            cargos_by_guia: Dict[str, List[dict]] = {}
            for (g, _ck), ev in last_event.items():
                # si el último estado es ELIMINAR => no se toma
                if ev.accion == "ELIMINAR":
                    continue

                cargos_by_guia.setdefault(g, []).append({
                    "cargo_id": ev.cargo_id.strip(),
                    "cargo": ev.cargo_nombre.strip() or ev.cargo_key,
                    "moneda": ev.moneda,
                    "monto": str(ev.monto),
                })

            # inyectar en guia_map
            for g, cargos in cargos_by_guia.items():
                if g not in guia_map:
                    guia_map[g] = {
                        "guia": g,
                        "contenedor": "",
                        "estado": "",
                        "fecha": None,
                        "fecha_cierre": None,
                        "ruta": "",
                        "monto_total": Decimal("0"),
                        "monto_flete": None,
                        "monto_extras": None,
                        "cargos": [],
                    }
                guia_map[g]["cargos"] = cargos

        else:
            logger.warning("FILS: no existe hoja 'Cargos Adicionales' (se continuará sin cargos).")

        try:
            wb.close()
        except Exception:
            pass

        # Salida: lista de dicts
        out = list(guia_map.values())

        # Normalizaciones extra útiles
        for r in out:
            # asegurar strings
            r["guia"] = str(r.get("guia", "")).strip()
            r["contenedor"] = str(r.get("contenedor", "")).strip().upper().replace("-", "").replace(" ", "")
            r["estado"] = str(r.get("estado", "")).strip().upper()
            r["ruta"] = str(r.get("ruta", "")).strip()

            # asegurar decimal string en monto_total (reconcile lo parsea con parse_money)
            mt = r.get("monto_total")
            if isinstance(mt, Decimal):
                r["monto_total"] = str(mt)
            else:
                r["monto_total"] = str(parse_money(mt))

        logger.info(f"FILS parse done: guias={len(out)}")
        return out
