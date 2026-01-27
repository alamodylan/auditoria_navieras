# app/parsers/one_facturacion.py

from __future__ import annotations

from typing import Dict, List, Iterable, Optional, Any
from openpyxl import load_workbook

from app.parsers.base import BaseParser
from app.parsers.normalization import normalize_guia, normalize_contenedor, normalize_amount
from app.utils.strings import upper_clean


class ONEFacturacionParser(BaseParser):
    """
    ONE puede variar mucho.
    Implementación streaming (openpyxl read_only) para soportar excels grandes en Render.
    Campos normalizados:
      guia, contenedor (si viene), total_naviera, ruta, sheet
    """

    SYNONYMS = {
        "guia": ["Guia", "Guía", "Documento", "No Documento", "Referencia", "Reference"],
        "contenedor": ["Contenedor", "Container", "CNTR"],
        "total": ["Total", "Monto", "Importe", "Amount", "Total Facturado"],
        "ruta": ["Ruta", "Servicio", "Service", "Tipo", "Servicio Facturado"],
    }

    def sniff(self, path: str) -> Dict:
        meta = {"errors": [], "warnings": []}
        try:
            wb = load_workbook(filename=path, read_only=True, data_only=True)
        except Exception as e:
            meta["errors"].append(f"ONE: no se pudo abrir el Excel: {e}")
            return meta

        try:
            sheets = [ws.title for ws in wb.worksheets]
            meta["sheets"] = sheets
            if not sheets:
                meta["errors"].append("ONE: el archivo no contiene hojas.")
                return meta

            ws = wb.worksheets[0]
            meta["sheet_used"] = ws.title

            # leer 1 fila de headers + 2 de muestra
            rows = []
            for r in ws.iter_rows(min_row=1, max_row=3, values_only=True):
                rows.append(list(r or []))

            if not rows:
                meta["errors"].append("ONE: hoja vacía.")
                return meta

            headers_raw = [str(x).strip() if x is not None else "" for x in rows[0]]
            meta["headers_preview"] = headers_raw[:30]

            idx = self._map_header_indices(headers_raw)

            # validar mínimos
            if idx.get("guia") is None:
                meta["errors"].append("ONE: no se encontró columna Guía/Documento/Referencia.")
            if idx.get("total") is None:
                meta["errors"].append("ONE: no se encontró columna Total/Monto/Amount.")

            # para debug visual
            meta["mapped_sample"] = {
                "guia": self._colname(headers_raw, idx.get("guia")),
                "contenedor": self._colname(headers_raw, idx.get("contenedor")),
                "total": self._colname(headers_raw, idx.get("total")),
                "ruta": self._colname(headers_raw, idx.get("ruta")),
            }

            return meta

        finally:
            try:
                wb.close()
            except Exception:
                pass

    def parse(self, path: str) -> Iterable[dict]:
        """
        Streaming generator: NO construye lista completa.
        """
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        try:
            if not wb.worksheets:
                return
                yield  # pragma: no cover

            ws = wb.worksheets[0]
            sheet_name = ws.title

            # headers en fila 1
            header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers_raw = [str(x).strip() if x is not None else "" for x in header_cells]
            idx = self._map_header_indices(headers_raw)

            guia_i = idx.get("guia")
            total_i = idx.get("total")
            if guia_i is None or total_i is None:
                raise ValueError("ONE: columnas mínimas no encontradas (guía/total).")

            cont_i = idx.get("contenedor")
            ruta_i = idx.get("ruta")

            for row in ws.iter_rows(min_row=2, values_only=True):
                guia = normalize_guia(self._cell(row, guia_i))
                if not guia:
                    continue

                yield {
                    "guia": guia,
                    "contenedor": normalize_contenedor(self._cell(row, cont_i)) if cont_i is not None else "",
                    "total_naviera": normalize_amount(self._cell(row, total_i)) or 0,
                    "ruta": str(self._cell(row, ruta_i) or "").strip() if ruta_i is not None else "",
                    "sheet": sheet_name,
                }

        finally:
            try:
                wb.close()
            except Exception:
                pass

    # -------------------------
    # Helpers
    # -------------------------

    def _cell(self, row: Any, idx: Optional[int]):
        if idx is None:
            return None
        if idx < 0:
            return None
        return row[idx] if idx < len(row) else None

    def _colname(self, headers_raw: List[str], idx: Optional[int]) -> str:
        if idx is None:
            return ""
        return headers_raw[idx] if idx < len(headers_raw) else ""

    def _map_header_indices(self, headers_raw: List[str]) -> Dict[str, Optional[int]]:
        """
        Mapea canonical -> índice de columna usando sinónimos.
        Comparación robusta con upper_clean (quita tildes, ¿?, etc.)
        """
        headers_norm = [upper_clean(h) for h in headers_raw]

        def find_idx(options: List[str]) -> Optional[int]:
            opts_norm = [upper_clean(o) for o in options]
            # match exact
            for j, hn in enumerate(headers_norm):
                if hn in opts_norm:
                    return j
            # match contains
            for j, hn in enumerate(headers_norm):
                for o in opts_norm:
                    if o and o in hn:
                        return j
            return None

        return {
            "guia": find_idx(self.SYNONYMS["guia"]),
            "contenedor": find_idx(self.SYNONYMS["contenedor"]),
            "total": find_idx(self.SYNONYMS["total"]),
            "ruta": find_idx(self.SYNONYMS["ruta"]),
        }