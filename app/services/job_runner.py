# app/services/job_runner.py

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from typing import Dict, List, Iterable, Optional
import unicodedata

from openpyxl import load_workbook

from app.extensions import db
from app.models import (
    Job,
    ResultSummary, ResultContainer, ResultCharge, ResultException, ResultKPI
)

from app.parsers.fils_auditoria import FILSAuditoriaParser
from app.parsers.cosco_facturacion import COSCOFacturacionParser
from app.parsers.one_facturacion import ONEFacturacionParser

from app.services.reconciliation import reconcile
from app.services.kpis import compute_kpis
from app.exporters.excel_export import export_job_to_excel

from app.utils.logging import get_logger

logger = get_logger("job_runner")

BATCH_SIZE = 1000


def _bulk_delete_job_results(job_id: int) -> None:
    ResultSummary.query.filter_by(job_id=job_id).delete(synchronize_session=False)
    ResultContainer.query.filter_by(job_id=job_id).delete(synchronize_session=False)
    ResultCharge.query.filter_by(job_id=job_id).delete(synchronize_session=False)
    ResultException.query.filter_by(job_id=job_id).delete(synchronize_session=False)
    ResultKPI.query.filter_by(job_id=job_id).delete(synchronize_session=False)
    db.session.commit()


def _bulk_insert(model, rows: List[dict]) -> None:
    if not rows:
        return
    db.session.bulk_insert_mappings(model, rows)
    db.session.commit()


def _strip_accents(s: str) -> str:
    if not s:
        return ""
    return "".join(
        c for c in unicodedata.normalize("NFKD", s)
        if not unicodedata.combining(c)
    )


def _norm_header(s: str) -> str:
    s = (s or "").strip().lower()
    s = _strip_accents(s)
    s = " ".join(s.split())
    return s


def _parse_fecha(value) -> Optional[datetime]:
    """
    Parsea fecha del FILS.
    Puede venir como datetime (openpyxl) o string tipo: 12/11/2024 11:20
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value

    s = str(value).strip()
    if not s:
        return None

    try:
        return datetime.strptime(s, "%d/%m/%Y %H:%M")
    except Exception:
        return None


def _detect_fils_header_row(fils_path: str) -> int:
    """
    Detecta si el header real está en fila 1 o 2.
    Regla: buscamos columnas clave (numero guia + monto tarifa)
    en las dos primeras filas.
    """
    try:
        wb = load_workbook(filename=fils_path, read_only=True, data_only=True)
        try:
            ws = wb.worksheets[0]
            row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            row2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)

            def row_has_keys(row) -> bool:
                if not row:
                    return False
                headers = [_norm_header(str(x) if x is not None else "") for x in row]
                has_guia = any("numero guia" in h for h in headers)
                has_monto = any("monto tarifa" in h for h in headers) or any("monto total" in h for h in headers) or any(h == "total" for h in headers)
                return has_guia and has_monto

            if row_has_keys(row1):
                return 1
            if row_has_keys(row2):
                return 2

            # fallback: por compatibilidad, asume fila 1
            return 1
        finally:
            try:
                wb.close()
            except Exception:
                pass
    except Exception:
        return 1


def run_job(job_id: int, money_tolerance: float, output_folder: str) -> Dict:
    job = Job.query.get(job_id)
    if not job:
        raise ValueError(f"Job no existe: {job_id}")

    job.mark_running()
    db.session.commit()

    try:
        # obtener paths
        files = {f.file_type.upper(): f for f in job.files}
        if "FILS" not in files:
            raise ValueError("Falta archivo FILS en el Job.")
        if job.naviera.upper() not in files:
            raise ValueError(f"Falta archivo de facturación {job.naviera} en el Job.")

        fils_path = files["FILS"].stored_path
        fact_path = files[job.naviera.upper()].stored_path

        tol = Decimal(str(money_tolerance))

        # ----------------------------
        # FILS streaming -> iterable dicts (alineado a headers reales)
        # ----------------------------
        fils_parser = FILSAuditoriaParser()
        header_row = _detect_fils_header_row(fils_path)
        logger.info(f"FILS detected header_row={header_row}")

        def iter_fils_dicts() -> Iterable[dict]:
            # OJO: headers vienen normalizados sin tildes (por fils_auditoria)
            guia_hints = ["numero guia"]
            cont_hints = ["contenedor"]
            estado_hints = ["estado"]
            fecha_hints = ["fecha"]
            ruta_hints = ["ruta"]
            monto_tarifa_hints = ["monto tarifa"]  # el total que vas a comparar

            def find_idx(headers: List[str], hints: List[str]) -> Optional[int]:
                for i, h in enumerate(headers):
                    hh = (h or "")
                    for hint in hints:
                        if hint in hh:
                            return i
                return None

            header_cached = None
            idx_map: Dict[str, Optional[int]] = {}

            for headers, row in fils_parser.iter_rows(fils_path, header_row=header_row):
                if not headers:
                    continue

                if header_cached is None:
                    header_cached = headers

                    idx_map["guia"] = find_idx(headers, guia_hints)
                    idx_map["contenedor"] = find_idx(headers, cont_hints)
                    idx_map["estado"] = find_idx(headers, estado_hints)
                    idx_map["fecha"] = find_idx(headers, fecha_hints)
                    idx_map["ruta"] = find_idx(headers, ruta_hints)
                    idx_map["monto_tarifa"] = find_idx(headers, monto_tarifa_hints)

                    if idx_map["guia"] is None:
                        raise ValueError("FILS: no se encontró columna 'Número Guía' (numero guia).")
                    if idx_map["monto_tarifa"] is None:
                        raise ValueError("FILS: no se encontró columna 'Monto Tarifa' (monto tarifa).")

                    logger.info(f"FILS idx_map={idx_map}")

                def cell(i: Optional[int]):
                    if i is None:
                        return None
                    return row[i] if i < len(row) else None

                guia = str(cell(idx_map["guia"]) or "").strip()
                if not guia:
                    continue

                cont = str(cell(idx_map["contenedor"]) or "").strip()
                cont = cont.upper().replace("-", "").replace(" ", "")

                estado = str(cell(idx_map["estado"]) or "").strip().upper()

                fecha = _parse_fecha(cell(idx_map["fecha"]))
                # clave para ordenar sin reventar: si None, usamos datetime.min
                fecha_sort = fecha if fecha is not None else datetime.min

                ruta = str(cell(idx_map["ruta"]) or "").strip()

                monto_tarifa = cell(idx_map["monto_tarifa"])

                yield {
                    "guia": guia,
                    "contenedor": cont,
                    "estado": estado,

                    # reconcile() usa esto para escoger "última cerrada"
                    "fecha": fecha_sort,
                    "fecha_cierre": fecha_sort,

                    # reconcile() toma monto_total primero
                    "monto_total": monto_tarifa,

                    "monto_flete": None,
                    "monto_extras": None,

                    "ruta": ruta,
                    "cargos": [],
                }

        # ----------------------------
        # Naviera parse (por ahora pandas -> lista)
        # ----------------------------
        if job.naviera.upper() == "COSCO":
            nav_parser = COSCOFacturacionParser()
        else:
            nav_parser = ONEFacturacionParser()

        nav_rows_iter: Iterable[dict] = nav_parser.parse(fact_path)

        # ----------------------------
        # Reconcile
        # ----------------------------
        resumen, det_cont, det_cargos, excs = reconcile(
            job.naviera,
            fils_rows=list(iter_fils_dicts()),   # 👈 reconcile actual asume lista en type hints y se reutiliza
            naviera_rows=list(nav_rows_iter),    # 👈 igual
            money_tolerance=tol
        )

        # ----------------------------
        # Persist results (limpiar previos)
        # ----------------------------
        _bulk_delete_job_results(job_id)

        # Summary bulk
        buf: List[dict] = []
        for r in resumen:
            buf.append({
                "job_id": job_id,
                "guia": r.guia,
                "estado": r.estado,
                "total_fils": r.total_fils,
                "total_naviera": r.total_naviera,
                "diferencia": r.diferencia,
                "ok": r.ok,
                "naviera": r.naviera,
                "fuente_naviera": r.fuente_naviera,
            })
            if len(buf) >= BATCH_SIZE:
                _bulk_insert(ResultSummary, buf)
                buf.clear()
        _bulk_insert(ResultSummary, buf)

        # Containers bulk
        buf = []
        for c in det_cont:
            buf.append({
                "job_id": job_id,
                "guia": str(c.get("guia", "")),
                "contenedor": str(c.get("contenedor", "")),
                "ruta": str(c.get("ruta", "")),
                "flete": c.get("flete") or 0,
                "extras": c.get("extras") or 0,
                "total": c.get("total") or 0,
                "naviera": str(c.get("naviera", "")),
            })
            if len(buf) >= BATCH_SIZE:
                _bulk_insert(ResultContainer, buf)
                buf.clear()
        _bulk_insert(ResultContainer, buf)

        # Charges bulk
        buf = []
        for ch in det_cargos:
            buf.append({
                "job_id": job_id,
                "guia": str(ch.get("guia", "")),
                "contenedor": str(ch.get("contenedor", "")),
                "tipo_cargo": str(ch.get("tipo_cargo", "CARGO")),
                "monto": ch.get("monto") or 0,
                "origen": str(ch.get("origen", "FILS")),
                "naviera": str(ch.get("naviera", "")),
            })
            if len(buf) >= BATCH_SIZE:
                _bulk_insert(ResultCharge, buf)
                buf.clear()
        _bulk_insert(ResultCharge, buf)

        # Exceptions bulk
        buf = []
        for e in excs:
            buf.append({
                "job_id": job_id,
                "tipo": e.tipo,
                "guia": e.guia,
                "contenedor": e.contenedor,
                "detalle": e.detalle,
                "severidad": e.severidad,
                "naviera": e.naviera,
            })
            if len(buf) >= BATCH_SIZE:
                _bulk_insert(ResultException, buf)
                buf.clear()
        _bulk_insert(ResultException, buf)

        # KPI
        resumen_dicts = [{
            "guia": r.guia,
            "ok": r.ok,
            "estado": r.estado,
            "total_fils": str(r.total_fils),
            "total_naviera": str(r.total_naviera),
            "diferencia": str(r.diferencia),
        } for r in resumen]

        kpi = compute_kpis(job.naviera.upper(), resumen_dicts)

        db.session.add(ResultKPI(
            job_id=job_id,
            naviera=kpi["naviera"],
            total_guias=kpi["total_guias"],
            guias_ok=kpi["guias_ok"],
            guias_diferencia=kpi["guias_diferencia"],
            guias_no_cerrada=kpi["guias_no_cerrada"],
            guias_solo_en_fils=kpi["guias_solo_en_fils"],
            guias_solo_en_naviera=kpi["guias_solo_en_naviera"],
            total_fils=kpi["total_fils"],
            total_naviera=kpi["total_naviera"],
            diferencia_global=kpi["diferencia_global"],
        ))
        db.session.commit()

        # Export
        export_path = export_job_to_excel(job_id=job_id, output_folder=output_folder)

        job.mark_done()
        db.session.commit()

        return {
            "job_id": job_id,
            "status": "DONE",
            "export_path": export_path,
            "kpi": kpi,
        }

    except Exception as e:
        job.mark_failed(e)
        db.session.commit()
        logger.exception(f"Job failed id={job_id}: {e}")
        return {"job_id": job_id, "status": "FAILED", "error": str(e)}

